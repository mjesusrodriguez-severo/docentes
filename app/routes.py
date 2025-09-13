import os
import traceback
from collections import defaultdict
from pathlib import Path

from sqlalchemy import func  # Asegúrate de tener este import
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort, current_app, send_file, make_response
from flask_login import login_required, current_user
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from sqlalchemy.orm import joinedload, selectinload
import unicodedata

from .models import Alumno, Grupo, Responsable, AlumnoResponsable, Usuario, Amonestacion, ReservaSala, \
    ReservaInformatica, Sustitucion, Ubicacion, Dispositivo, dispositivos_reservados, Incidencia, \
    IncidenciaMantenimiento, InformeAlumno, InformeFaltas
from . import db
from datetime import datetime, date, timedelta
from pytz import timezone
from flask_mail import Message
from . import mail
import io

import requests
from requests.auth import HTTPBasicAuth
from werkzeug.utils import secure_filename

from .utils.decoradores import rol_requerido
from .utils.drive import subir_archivo_a_drive, crear_carpeta_sustitucion
from .utils.google_auth import build_drive_service
from .utils.reservas import render_calendario_espacio, enviar_correo_reserva_espacio
from .utils.sms import enviar_sms_esendex

import pandas as pd

from flask import send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo

main_bp = Blueprint('main', __name__)
# Hora actual en España
fecha_espana = datetime.now(ZoneInfo("Europe/Madrid"))

@main_bp.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))
    return render_template("login.html")

@main_bp.route("/dashboard")
@login_required
def dashboard():
    hoy = date.today()

    # Sustituciones
    if current_user.rol == "jefatura":
        mis_sustituciones = Sustitucion.query.filter(Sustitucion.fecha >= hoy).order_by(Sustitucion.fecha).all()
    else:
        mis_sustituciones = Sustitucion.query.filter(
            Sustitucion.sustituto_id == current_user.id,
            Sustitucion.fecha >= hoy
        ).order_by(Sustitucion.fecha).all()

    # Reservas
    reservas_sala = ReservaSala.query.filter(
        ReservaSala.usuario_id == current_user.id,
        ReservaSala.fecha >= hoy
    ).order_by(ReservaSala.fecha).all()

    reservas_tic = ReservaInformatica.query.filter(
        ReservaInformatica.usuario_id == current_user.id,
        ReservaInformatica.fecha >= hoy
    ).order_by(ReservaInformatica.fecha).all()

    # Amonestaciones
    grupo_tutoria = Grupo.query.filter_by(tutor_id=current_user.id).first()
    es_tutor = grupo_tutoria is not None

    if current_user.rol == "jefatura":
        amonestaciones_personales = Amonestacion.query.order_by(Amonestacion.fecha.desc()).limit(5).all()
        amonestaciones_tutoria = []
    else:
        amonestaciones_personales = Amonestacion.query.filter(
            Amonestacion.profesor_id == current_user.id
        ).order_by(Amonestacion.fecha.desc()).limit(5).all()

        if es_tutor:
            # IDs de alumnos del grupo que tutoriza
            alumnos_tutoria_ids = [a.id for a in Alumno.query.filter_by(grupo_id=grupo_tutoria.id).all()]
            amonestaciones_tutoria = Amonestacion.query.filter(
                Amonestacion.alumno_id.in_(alumnos_tutoria_ids)
            ).order_by(Amonestacion.fecha.desc()).limit(5).all()
        else:
            amonestaciones_tutoria = []

    return render_template("dashboard.html",
                           sustituciones=mis_sustituciones,
                           reservas_sala=reservas_sala,
                           reservas_tic=reservas_tic,
                           amonestaciones_personales=amonestaciones_personales,
                           amonestaciones_tutoria=amonestaciones_tutoria,
                           num_sustituciones=len(mis_sustituciones),
                           num_reservas_sala=len(reservas_sala),
                           num_reservas_tic=len(reservas_tic),
                           num_amonestaciones=len(amonestaciones_personales) + len(amonestaciones_tutoria),
                           es_tutor=es_tutor,
                           nombre_grupo_tutoria=grupo_tutoria.nombre if es_tutor else None)

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                            RUTAS DE USUARIOS                           ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/usuarios")
@login_required
@rol_requerido("tic", "jefatura")
def listar_usuarios():
    usuarios = Usuario.query.all()
    return render_template("usuarios/listar_usuarios.html", usuarios=usuarios)

@main_bp.route("/usuarios/nuevo", methods=["GET", "POST"])
@login_required
@rol_requerido("tic")
def nuevo_usuario():
    if request.method == "POST":
        email = request.form["email"]
        nombre = request.form["nombre"]
        rol = request.form["rol"]
        telefono = request.form["telefono"]

        nuevo = Usuario(email=email, nombre=nombre, rol=rol, telefono=telefono)
        db.session.add(nuevo)
        db.session.commit()
        flash("Usuario añadido correctamente.")
        return redirect(url_for("main.listar_usuarios"))

    return render_template("usuarios/nuevo_usuario.html")

@main_bp.route("/usuarios/editar/<int:usuario_id>", methods=["GET", "POST"])
@login_required
@rol_requerido("tic", "jefatura")
def editar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)

    if request.method == "POST":
        usuario.nombre = request.form["nombre"]
        usuario.rol = request.form["rol"]
        usuario.telefono = request.form["telefono"]
        usuario.enlace_materiales = request.form.get("enlace_materiales")

        db.session.commit()
        flash("Usuario actualizado correctamente.")
        return redirect(url_for("main.listar_usuarios"))

    return render_template("usuarios/editar_usuario.html", usuario=usuario)

@main_bp.route("/usuarios/eliminar/<int:usuario_id>", methods=["POST"])
@login_required
@rol_requerido("tic")
def eliminar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    db.session.delete(usuario)
    db.session.commit()
    flash("Usuario eliminado correctamente.")
    return redirect(url_for("main.listar_usuarios"))

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                            RUTAS DE ALUMNOS                            ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/alumnos")
@main_bp.route("/alumnos/<int:grupo_id>")
@login_required
def listar_alumnos(grupo_id=None):
    from .models import Grupo, Alumno

    grupos = Grupo.query.order_by(Grupo.orden).all()
    grupo_seleccionado = None
    alumnos = []

    if grupo_id:
        grupo_seleccionado = Grupo.query.get_or_404(grupo_id)
        alumnos = Alumno.query.filter_by(grupo_id=grupo_id).order_by(Alumno.apellidos, Alumno.nombre).all()

    return render_template(
        "alumnos/alumnos_por_grupo.html",
        grupos=grupos,
        grupo_seleccionado=grupo_seleccionado,
        alumnos=alumnos
    )

@main_bp.route("/alumno/<int:alumno_id>")
@login_required
def ver_alumno(alumno_id):
    from .models import Alumno, AlumnoResponsable

    alumno = Alumno.query.get_or_404(alumno_id)

    # Obtener responsable principal
    asoc_principal = AlumnoResponsable.query.filter_by(alumno_id=alumno_id, principal=True).first()
    id_responsable_principal = asoc_principal.responsable_id if asoc_principal else None

    return render_template(
        "alumnos/ver_alumno.html",
        alumno=alumno,
        id_responsable_principal=id_responsable_principal
    )


@main_bp.route("/alumno/<int:alumno_id>/eliminar", methods=["POST"])
@login_required
def eliminar_alumno(alumno_id):
    from .models import Alumno, Responsable, db

    alumno = Alumno.query.get_or_404(alumno_id)

    # Guardar los responsables asociados antes de eliminar el alumno
    responsables_asociados = list(alumno.responsables)

    grupo_id = alumno.grupo_id  # Guardamos grupo antes de borrar

    # Eliminar al alumno
    db.session.delete(alumno)
    db.session.commit()

    # Comprobar si alguno de los responsables ya no tiene alumnos
    for responsable in responsables_asociados:
        if not responsable.alumnos:  # No tiene alumnos asociados
            db.session.delete(responsable)

    db.session.commit()

    return redirect(url_for("main.listar_alumnos", grupo_id=grupo_id))

@main_bp.route('/grupo/<int:grupo_id>/nuevo_alumno', methods=['GET', 'POST'])
@login_required
def nuevo_alumno(grupo_id):
    grupo = Grupo.query.get_or_404(grupo_id)

    if request.method == 'POST':
        # Crear nuevo alumno con todos los campos
        nuevo_alumno = Alumno(
            nombre=request.form['nombre'],
            apellidos=request.form['apellidos'],
            identificacion=request.form['identificacion'],
            direccion=request.form.get('direccion'),
            telefono=request.form.get('telefono'),
            email=request.form.get('email'),
            observaciones=request.form.get('observaciones'),
            grupo_id=grupo.id
        )

        db.session.add(nuevo_alumno)
        db.session.flush()  # Obtener ID del alumno antes de commit

        # ----------------------------
        # Asociar responsables existentes (select2)
        # ----------------------------
        responsable_ids = request.form.getlist('responsable_ids')
        for rid in responsable_ids:
            responsable = Responsable.query.get(int(rid))
            if responsable:
                nuevo_alumno.responsables.append(responsable)

        # ----------------------------
        # Añadir nuevos responsables
        # ----------------------------
        index = 0
        while True:
            nombre_key = f'nuevo_nombre_{index}'
            telefono_key = f'nuevo_telefono_{index}'
            email_key = f'nuevo_email_{index}'

            if nombre_key not in request.form:
                break  # No hay más nuevos responsables

            nombre = request.form[nombre_key].strip()
            telefono = request.form[telefono_key].strip()
            email = request.form[email_key].strip()

            if nombre:
                responsable = Responsable(
                    nombre=nombre,
                    telefono=telefono,
                    email=email
                )
                db.session.add(responsable)
                db.session.flush()  # Para obtener el ID

                nuevo_alumno.responsables.append(responsable)

                # Ver si es el responsable principal
                if request.form.get('principal_id') == f'nuevo_{index}':
                    responsable.principal = True

            index += 1

        db.session.commit()
        flash('Alumno registrado correctamente.', 'success')
        return redirect(url_for('main.listar_alumnos', grupo_id=grupo.id))

    return render_template('alumnos/nuevo_alumno.html', grupo=grupo)

@main_bp.route('/buscar_responsables')
@login_required
def buscar_responsables():
    termino = request.args.get('q', '').strip()

    if not termino:
        return jsonify([])

    resultados = Responsable.query.filter(
        (Responsable.nombre.ilike(f"%{termino}%")) |
        (Responsable.telefono.ilike(f"%{termino}%"))
    ).limit(10).all()

    data = [
        {
            'id': responsable.id,
            'text': f"{responsable.nombre} ({responsable.telefono})"
        }
        for responsable in resultados
    ]

    return jsonify(data)

@main_bp.route('/alumnos/<int:alumno_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_alumno(alumno_id):
    alumno = Alumno.query.get_or_404(alumno_id)
    grupo = alumno.grupo
    todos_los_responsables = Responsable.query.all()

    asociaciones_actuales = alumno.alumno_responsable_asociaciones
    id_responsable_principal = None
    for asoc in asociaciones_actuales:
        if asoc.principal:
            id_responsable_principal = asoc.responsable_id
            break

    if request.method == 'POST':
        # 1. Actualizar datos del alumno
        alumno.nombre = request.form['nombre']
        alumno.apellidos = request.form['apellidos']
        alumno.identificacion = request.form.get('identificacion')
        alumno.direccion = request.form.get('direccion')
        alumno.telefono = request.form.get('telefono')
        alumno.email = request.form.get('email')
        alumno.observaciones = request.form.get('observaciones')
        alumno.grupo_id = request.form.get('grupo_id')

        # 2. Responsable principal seleccionado
        id_nuevo_principal = request.form.get('principal')  # puede ser un ID o "nuevo"

        # 3. Eliminar responsables seleccionados
        ids_eliminados = request.form.getlist('eliminar_responsables')

        for asoc in asociaciones_actuales:
            responsable = asoc.responsable
            rid = str(responsable.id)

            if rid in ids_eliminados:
                db.session.delete(asoc)
                continue

            # Actualizar nombre y teléfono del responsable
            responsable.nombre = request.form.get(f'nombre_responsable_{rid}', responsable.nombre)
            responsable.telefono = request.form.get(f'telefono_responsable_{rid}', responsable.telefono)

            # Marcar si es principal
            asoc.principal = (rid == id_nuevo_principal)

        # 4. Añadir responsable existente desde el desplegable
        nuevo_id = request.form.get("nuevo_responsable_id")
        if nuevo_id:
            nuevo_rid = int(nuevo_id)
            if nuevo_rid not in [asoc.responsable_id for asoc in asociaciones_actuales]:
                responsable = Responsable.query.get(nuevo_rid)
                if responsable:
                    nueva_asoc = AlumnoResponsable(
                        alumno_id=alumno.id,
                        responsable_id=nuevo_rid,
                        principal=(nuevo_id == id_nuevo_principal)
                    )
                    db.session.add(nueva_asoc)

        # 5. Crear nuevo responsable si se ha rellenado el formulario
        nuevo_nombre = request.form.get("nuevo_nombre")
        nuevo_telefono = request.form.get("nuevo_telefono")
        if nuevo_nombre and nuevo_telefono:
            nuevo_responsable = Responsable(nombre=nuevo_nombre, telefono=nuevo_telefono)
            db.session.add(nuevo_responsable)
            db.session.flush()  # Para obtener su ID antes del commit

            nueva_asoc = AlumnoResponsable(
                alumno_id=alumno.id,
                responsable_id=nuevo_responsable.id,
                principal=(id_nuevo_principal == "nuevo")
            )
            db.session.add(nueva_asoc)

        db.session.commit()
        flash("Alumno actualizado correctamente", "success")
        return redirect(url_for('main.listar_alumnos', grupo_id=alumno.grupo_id))

    # Para el formulario GET
    responsables_actuales = [
        {
            "asoc": asoc,
            "responsable": asoc.responsable
        } for asoc in asociaciones_actuales
    ]

    responsables_ids = [asoc.responsable_id for asoc in asociaciones_actuales]
    responsables_no_asociados = [r for r in todos_los_responsables if r.id not in responsables_ids]

    return render_template(
        'alumnos/editar_alumno.html',
        alumno=alumno,
        grupo=grupo,
        grupos=Grupo.query.all(),
        responsables_actuales=responsables_actuales,
        id_responsable_principal=id_responsable_principal,
        responsables_no_asociados=responsables_no_asociados
    )

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                            RUTAS DE GRUPOS                             ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/grupos")
@login_required
@rol_requerido("tic")
def listar_grupos():
    from .models import Grupo
    grupos = Grupo.query.order_by(Grupo.orden).all()
    return render_template("grupos/listar_grupos.html", grupos=grupos)

@main_bp.route("/grupo/<int:grupo_id>/editar", methods=["GET", "POST"])
@login_required
@rol_requerido("tic")
def editar_grupo(grupo_id):
    from .models import Grupo, Usuario, db

    grupo = Grupo.query.get_or_404(grupo_id)

    # Todos los docentes disponibles para seleccionar como profesores del grupo
    profesores = Usuario.query.all()
    tutores_posibles = profesores  # Por si quieres usarlos en la plantilla

    if request.method == "POST":
        grupo.nombre = request.form["nombre"]
        grupo.periodo = request.form["periodo"]
        grupo.orden = request.form.get("orden") or None

        # ---------- Tutor ----------
        tutor_id = request.form.get("tutor_id")
        if tutor_id:
            tutor = Usuario.query.get(int(tutor_id))
            if tutor and tutor.rol != "tutor":
                tutor.rol = "tutor"  # cambia el rol automáticamente si no lo tiene
            grupo.tutor = tutor
        else:
            grupo.tutor = None

        # ---------- Profesores asignados ----------
        ids_profesores = request.form.getlist("profesores_ids")
        grupo.profesores = Usuario.query.filter(Usuario.id.in_(ids_profesores)).all()

        db.session.commit()
        flash("Grupo actualizado correctamente.", "success")
        return redirect(url_for("main.listar_grupos"))

    return render_template(
        "grupos/editar_grupo.html",
        grupo=grupo,
        profesores=profesores,
        tutores_posibles=tutores_posibles
    )

    tutores_posibles = Usuario.query.filter_by(archivado=False).all()
    return render_template("grupos/editar_grupo.html", grupo=grupo, tutores_posibles=tutores_posibles, profesores=profesores)

@main_bp.route('/grupo/<int:grupo_id>/alumnos')
@login_required
def obtener_alumnos(grupo_id):
    alumnos = Alumno.query.filter_by(grupo_id=grupo_id).order_by(Alumno.apellidos).all()
    alumnos_json = [{'id': a.id, 'nombre': a.nombre + " " + a.apellidos} for a in alumnos]
    return jsonify(alumnos_json)

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                        RUTAS DE AMONESTACIONES                         ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/amonestaciones", methods=["GET", "POST"])
@login_required
def crear_amonestacion():
    if request.method == "POST":
        alumno_id = request.form.get("alumno_id")
        motivo = request.form.get("motivo")
        descripcion = request.form.get("descripcion")  # 👈 Añadir esto

        amonestacion = Amonestacion(
            alumno_id=alumno_id,
            profesor_id=current_user.id,
            motivo=motivo,
            descripcion=descripcion,
            fecha=fecha_espana
        )
        db.session.add(amonestacion)
        db.session.commit()

        alumno = Alumno.query.get(alumno_id)
        grupo = Grupo.query.get(alumno.grupo_id)
        tutor = Usuario.query.get(grupo.tutor_id)

        if tutor and tutor.telefono:
            mensaje = (
                f"📢 Jefatura de estudios: "
                f"Se ha registrado una nueva amonestación para {alumno.nombre}, del grupo {grupo.nombre}. "
                f"Motivo: {motivo}. Por favor, revisa la plataforma."
            )

            ok, respuesta = enviar_sms_esendex(tutor.telefono, mensaje)
            if not ok:
                flash("Error al enviar SMS al tutor.", "danger")

        return redirect(url_for("main.crear_amonestacion"))

    # GET
    grupos = Grupo.query.order_by(Grupo.orden).all()
    page = request.args.get("page", 1, type=int)
    if current_user.rol in ["jefatura", "tic"]:
        amonestaciones_query = Amonestacion.query
    elif current_user.rol == "tutor":
        grupo_tutoria = Grupo.query.filter_by(tutor_id=current_user.id).first()
        if grupo_tutoria:
            amonestaciones_query = Amonestacion.query.filter(
                (Amonestacion.profesor_id == current_user.id) |
                (Amonestacion.alumno.has(grupo_id=grupo_tutoria.id))
            )
        else:
            amonestaciones_query = Amonestacion.query.filter_by(profesor_id=current_user.id)
    else:
        amonestaciones_query = Amonestacion.query.filter_by(profesor_id=current_user.id)
    # Aplicar orden y paginación
    amonestaciones = amonestaciones_query.order_by(Amonestacion.fecha.desc()).paginate(page=page, per_page=10)

    return render_template("amonestaciones.html", grupos=grupos, amonestaciones=amonestaciones, es_tutor=current_user.rol == "tutor", grupo_tutoria=grupo_tutoria if current_user.rol == "tutor" else None)

@main_bp.route("/enviar_amonestacion", methods=["POST"])
@login_required
@rol_requerido("jefatura")
def enviar_amonestacion():
    amonestacion_id = request.form.get("id")
    responsable_id = request.form.get("responsable_id")

    amonestacion = Amonestacion.query.get_or_404(amonestacion_id)
    responsable = Responsable.query.get_or_404(responsable_id)

    if amonestacion.estado != "aceptada":
        flash("La amonestación debe estar aceptada antes de enviarla", "warning")
        return redirect(url_for("main.crear_amonestacion"))

    mensaje = f"Amonestación para {amonestacion.alumno.nombre} {amonestacion.alumno.apellidos}:\n"
    mensaje += f"Motivo: {amonestacion.motivo}\n"
    mensaje += f"Descripción: {amonestacion.descripcion}\n"
    mensaje += f"Fecha: {amonestacion.fecha.strftime('%d/%m/%Y %H:%M')}"

    # Mostrar la URL de WhatsApp en consola (puedes integrarla en HTML si quieres)
    from urllib.parse import quote
    mensaje_whatsapp = quote(mensaje)
    url = f"https://wa.me/{responsable.telefono}?text={mensaje_whatsapp}"
    print(f"[DEBUG] WhatsApp URL: {url}")

    flash(f"Amonestación preparada para envío a {responsable.nombre}.", "success")
    return redirect(url_for("main.crear_amonestacion"))

@main_bp.route("/revisar_amonestacion", methods=["POST"])
@login_required
def revisar_amonestacion():
    if not current_user.es_jefatura:
        abort(403)

    id = request.form.get("id")
    accion = request.form.get("accion")

    amonestacion = Amonestacion.query.get_or_404(id)
    if accion == "aceptar":
        amonestacion.estado = "aceptada"
    elif accion == "rechazar":
        amonestacion.estado = "rechazada"

    db.session.commit()
    flash("Amonestación revisada", "success")
    return redirect(url_for("main.crear_amonestacion"))

def enviar_sms_esendex_php_style(telefono, mensaje):
    user = "manolojimenez86@gmail.com"
    password = "b072431a128749aca763"
    account_ref = "EX0322259"
    remitente = "SeveroOchoa"

    mensaje += " | Ver en: https://severoochoa.es/sustituciones"

    xml_data = f"""
    <messages>
        <accountreference>{account_ref}</accountreference>
        <message>
            <to>{telefono}</to>
            <body>{mensaje}</body>
            <type>SMS</type>
            <originator>{remitente}</originator>
        </message>
    </messages>
    """.strip()

    headers = {"Content-Type": "text/xml"}

    response = requests.post(
        "https://api.esendex.com/v1.0/messagedispatcher",
        headers=headers,
        data=xml_data.encode("utf-8"),
        auth=HTTPBasicAuth(user, password)
    )

    return response.status_code in [200, 201], response.text

@main_bp.route("/enviar_sms/<int:amonestacion_id>", methods=["POST"])
@login_required
def enviar_sms_amonestacion(amonestacion_id):
    amonestacion_id = request.form.get("id")
    responsable_id = request.form.get("responsable_id")

    amon = Amonestacion.query.get_or_404(amonestacion_id)
    responsable = Responsable.query.get_or_404(responsable_id)

    telefono = responsable.telefono

    ok, respuesta = enviar_sms_amonestacion(telefono, amon)

    if ok:
        amon.enviado_responsables = True
        amon.fecha_envio_sms = fecha_espana
        db.session.commit()
        flash("SMS enviado correctamente", "success")
    else:
        flash(f"Error al enviar SMS: {respuesta}", "danger")

    return redirect(url_for("main.crear_amonestacion", amonestacion_id=amonestacion_id))

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                            RUTAS DE RESERVAS                           ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/reservas/sala-reuniones")
@login_required
def ver_calendario_sala_reuniones():
    return render_calendario_espacio(
        nombre_espacio= "sala_reuniones",
        plantilla="reservas/sala_reuniones.html",
        nombre_visible="Sala de reuniones")

@main_bp.route("/reservas/aula-taller")
@login_required
def ver_calendario_aula_taller():
    franjas_bloqueadas = {
        'Monday': ["12:40-13:35", "13:35-14:30"],
        'Tuesday': ["08:30-09:25", "09:25-10:25", "11:45-12:40", "13:35-14:30"],
        'Wednesday': ["08:30-09:25", "10:25-11:15", "11:45-12:40", "12:40-13:35"],
        'Thursday': ["08:30-09:25", "12:40-13:35", "13:35-14:30"],
        'Friday': ["08:30-09:25", "09:25-10:25", "12:40-13:35", "13:35-14:30"]
    }
    return render_calendario_espacio(
        franjas_bloqueadas=franjas_bloqueadas,
        nombre_espacio="aula_taller",
        plantilla="reservas/aula_taller.html",
        nombre_visible="Aula Taller"
    )

@main_bp.route("/reservas/departamento-taquillas")
@login_required
def ver_calendario_departamento_taquillas():
    return render_calendario_espacio(
        nombre_espacio="departamento_taquillas",
        plantilla="reservas/reuniones_taquillas.html",
        nombre_visible="Departamento Taquillas"
    )

@main_bp.route("/reservas/aula-laboratorio")
@login_required
def ver_calendario_aula_laboratorio():
    return render_calendario_espacio(
        nombre_espacio="aula_laboratorio",
        plantilla="reservas/laboratorio.html",
        nombre_visible="Laboratorio"
    )

@main_bp.route("/reservas/aula-digital")
@login_required
def ver_calendario_aula_digital():
    franjas_bloqueadas = {
        "Monday":    [],
        "Tuesday":   ["08:30-09:25", "09:25-10:25"],  # 1ª y 2ª hora
        "Wednesday": ["10:25-11:15", "11:45-12:40", "12:40-13:35"],  # 3ª, 4ª y 5ª hora
        "Thursday":  ["12:40-13:35", "13:35-14:30"],  # 5ª y 6ª hora
        "Friday":    ["08:30-09:25", "13:35-14:30"]   # 1ª y 6ª hora
    }
    return render_calendario_espacio(
        franjas_bloqueadas = franjas_bloqueadas,
        nombre_espacio="aula_digital",
        plantilla="reservas/aula_digital.html",
        nombre_visible="Aula Digital"
    )

@main_bp.route("/reservas/biblioteca")
@login_required
def ver_calendario_biblioteca():
    return render_calendario_espacio(
        nombre_espacio="biblioteca",
        plantilla="reservas/biblioteca.html",
        nombre_visible="Biblioteca"
    )

@main_bp.route("/reservas/<espacio>/reservar", methods=["POST"])
@login_required
def reservar_espacio(espacio):
    fecha = request.form.get("fecha")
    franja = request.form.get("franja")

    ya_reservada = ReservaSala.query.filter_by(
        fecha=fecha,
        franja_horaria=franja,
        espacio=espacio
    ).first()

    if ya_reservada:
        flash("Esta franja ya está reservada.", "danger")
    else:
        reserva = ReservaSala(
            usuario_id=current_user.id,
            fecha=fecha,
            franja_horaria=franja,
            espacio=espacio
        )
        db.session.add(reserva)
        db.session.commit()
        flash("Reserva realizada con éxito.", "success")

        enviar_correo_reserva_espacio(reserva, espacio, current_user)

    return redirect(url_for(f"main.ver_calendario_{espacio}"))

@main_bp.route("/reservas/<espacio>/cancelar", methods=["POST"])
@login_required
def cancelar_reserva_espacio(espacio):
    reserva_id = request.form.get("reserva_id")
    reserva = ReservaSala.query.get_or_404(reserva_id)

    if reserva.usuario_id != current_user.id:
        flash("No tienes permiso para cancelar esta reserva.", "danger")
    else:
        db.session.delete(reserva)
        db.session.commit()
        flash("Reserva cancelada correctamente.", "success")

    return redirect(url_for(f"main.ver_calendario_{espacio}"))

@main_bp.route("/reservar-material", methods=["GET", "POST"])
@login_required
def reservas_material():
    if request.method == "POST":
        reserva = ReservaInformatica(
            usuario_id=current_user.id,
            fecha=request.form["fecha"],
            franja=request.form["franja"],
            grupo_id=request.form["grupo_id"],
            cantidad=request.form["cantidad"],
            observaciones=request.form["observaciones"]
        )
        db.session.add(reserva)
        db.session.commit()
        flash("Reserva realizada correctamente", "success")
        return redirect(url_for("main.reservas_material"))

    # GET
    hoy = date.today()
    dias_semana_actual = [hoy + timedelta(days=i) for i in range(7) if (hoy + timedelta(days=i)).weekday() < 5]
    dias_semana_siguiente = [hoy + timedelta(days=i + 7) for i in range(7) if
                             (hoy + timedelta(days=i + 7)).weekday() < 5]
    franjas_horarias = [
        "08:30-09:25",
        "09:25-10:25",
        "10:25-11:15",
        "11:45-12:40",
        "12:40-13:35",
        "13:35-14:30",
    ]
    reservas = ReservaInformatica.query.filter(ReservaInformatica.estado != "CANCELADA").all()
    reservas_dict = defaultdict(list)
    for r in reservas:
        reservas_dict[(r.fecha, r.franja_horaria)].append(r)

    usuarios = {u.id: u.nombre for u in Usuario.query.all()}
    grupos = Grupo.query.order_by(Grupo.orden).all()

    # Paginación de reservas del usuario actual
    page = request.args.get('page', 1, type=int)
    if current_user.rol == "tic":
        # Mostrar todas las reservas si eres TIC
        reservas_usuario = ReservaInformatica.query \
            .order_by(ReservaInformatica.fecha.desc(), ReservaInformatica.franja_horaria.desc()) \
            .paginate(page=page, per_page=5)
    else:
        # Solo mostrar las del propio usuario
        reservas_usuario = ReservaInformatica.query.filter_by(usuario_id=current_user.id) \
            .order_by(ReservaInformatica.fecha.desc(), ReservaInformatica.franja_horaria.desc()) \
            .paginate(page=page, per_page=5)

    # DEBUG: mostrar tipo de cada valor en el diccionario 'reservas'
    for clave, valor in reservas_dict.items():
        print(f"Clave: {clave}, Tipo de valor: {type(valor)}")

    return render_template("reserva_material/calendario.html",
                           dias_semana_actual=dias_semana_actual,
                           dias_semana_siguiente=dias_semana_siguiente,
                           franjas_horarias=franjas_horarias,
                           reservas=reservas_dict,
                           usuarios=usuarios,
                           grupos=grupos,
                           reservas_usuario=reservas_usuario,
                           current_date=date.today(),
                           dias_es={"Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
                                    "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado",
                                    "Sunday": "Domingo"})

@main_bp.route('/reserva-material', methods=['GET', 'POST'])
@login_required
def formulario_reserva_material():
    fecha = request.args.get('fecha')
    franja = request.args.get('franja')

    grupos = Grupo.query.order_by(Grupo.orden).all()  # Ordenado por orden

    if request.method == 'POST':
        nueva_reserva = ReservaInformatica(
            usuario_id=current_user.id,
            grupo_id=request.form['grupo_id'],
            tipo_equipo=request.form.get('tipo_equipo', 'portátil'),  # ← opcional con valor por defecto
            cantidad=request.form['cantidad'],
            observaciones=request.form.get('observaciones'),
            fecha=request.form['fecha'],
            franja_horaria=request.form['franja'],
            estado='PENDIENTE'
        )
        db.session.add(nueva_reserva)
        db.session.commit()

        # Enviar email
        try:
            msg = Message(
                subject="Nueva solicitud de reserva",
                sender=("Panel de Docentes", current_app.config["MAIL_USERNAME"]),
                recipients=["mjesusrodriguez@severoochoa.es"],  # <-- Cámbialo por el tuyo
                html=f"""
                <h2 style="color:#2c3e50;">Nueva reserva de material realizada</h2>
                <p><strong>📅 Fecha:</strong> {nueva_reserva.fecha}</p>
                <p><strong>🕒 Franja horaria:</strong> {nueva_reserva.franja_horaria}</p>
                <p><strong>👩‍🏫 Profesor:</strong> {current_user.nombre}</p>
                <p><strong>👨‍👩‍👧 Grupo:</strong> {nueva_reserva.grupo.nombre}</p>
                <p><strong>💻 Tipo de equipo:</strong> {nueva_reserva.tipo_equipo.capitalize()}</p>
                <p><strong>🔢 Cantidad:</strong> {nueva_reserva.cantidad}</p>
                <p><strong>📝 Observaciones:</strong> {nueva_reserva.observaciones or "Ninguna"}</p>
                <p><strong>📌 Estado:</strong> <span style="color:orange;"><strong>{nueva_reserva.estado}</strong></span></p>
                <hr>
                <p style="font-size:0.9em; color:#888;">Este mensaje ha sido generado automáticamente por el sistema de reservas del centro.</p>
                """
            )
            mail.send(msg)
        except Exception as e:
            print("ERROR EN ENVÍO DE EMAIL:", e)
            flash("Reserva realizada, pero no se pudo enviar el correo: " + str(e), "warning")
        else:
            flash("Reserva realizada correctamente y correo enviado.", "success")

        return redirect(url_for('main.reservas_material'))

    return render_template('reserva_material/formulario_reserva.html',
                           fecha=fecha,
                           franja=franja,
                           grupos=grupos)

@main_bp.route("/cancelar-reserva-material", methods=["POST"])
@login_required
def cancelar_reserva_material():
    reserva_id = request.form.get("reserva_id")

    reserva = ReservaInformatica.query.get(reserva_id)
    if not reserva:
        flash("Reserva no encontrada", "danger")
        return redirect(url_for("main.reservas_material"))

    if reserva.usuario_id != current_user.id:
        flash("No tienes permiso para cancelar esta reserva", "danger")
        return redirect(url_for("main.reservas_material"))

    reserva.estado = "CANCELADA"
    db.session.commit()
    flash("Reserva cancelada correctamente", "success")
    return redirect(url_for("main.reservas_material"))

@main_bp.route("/descargar-reserva/<int:reserva_id>")
@login_required
def descargar_reserva(reserva_id):
    reserva = ReservaInformatica.query.get_or_404(reserva_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                            leftMargin=2 * cm, rightMargin=2 * cm)

    estilos = getSampleStyleSheet()
    normal = estilos["Normal"]
    normal.fontSize = 11

    titulo = estilos["Title"]
    titulo.fontSize = 16
    titulo.alignment = 1  # Center

    subtitulo = estilos["Normal"]
    subtitulo.fontSize = 12
    subtitulo.alignment = 1  # Center

    negrita = estilos["Heading4"]
    negrita.fontSize = 12

    elementos = []

    # Cabecera con logo a la izquierda y bloque centrado (título + subtítulo) a la derecha
    ruta_logo = os.path.join(current_app.root_path, 'static', 'img', 'logo.JPG')
    logo = Image(ruta_logo, width=3.5 * cm, height=3.5 * cm)

    # Texto del título y subtítulo dentro de una celda
    bloque_texto = [
        Paragraph("<b>PRÉSTAMO DE EQUIPOS INFORMÁTICOS</b>", titulo),
        Paragraph("C.C. Severo Ochoa", subtitulo)
    ]

    cabecera = Table(
        [[logo, bloque_texto]],
        colWidths=[3.5 * cm, 12.5 * cm]
    )
    cabecera.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    elementos.append(cabecera)
    elementos.append(Spacer(1, 0.5 * cm))

    # Datos del préstamo
    datos = [
        f"<b>Docente:</b> {reserva.usuario.nombre}",
        f"<b>Fecha:</b> {reserva.fecha.strftime('%d/%m/%Y')}",
        f"<b>Franja horaria:</b> {reserva.franja_horaria}",
        f"<b>Grupo:</b> {reserva.grupo.nombre}"
    ]
    for d in datos:
        elementos.append(Paragraph(d, negrita))

    elementos.append(Spacer(1, 0.5 * cm))

    # Tabla con 25 filas vacías
    encabezado = ["Nº equipo", "Nombre del alumno", "Observaciones"]
    filas = [encabezado] + [["", "", ""] for _ in range(25)]

    tabla = Table(filas, colWidths=[3 * cm, 8 * cm, 5 * cm])
    gris_claro = colors.Color(0.9, 0.9, 0.9)  # RGB entre 0 y 1

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), gris_claro),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    filename = f"solicitud_prestamo_{reserva.id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

@main_bp.route('/subir-hoja/<int:reserva_id>', methods=['POST'])
@login_required
def subir_hoja(reserva_id):
    reserva = ReservaInformatica.query.get_or_404(reserva_id)
    archivo = request.files['archivo']
    drive_service = build_drive_service()

    if archivo:
        # Guardar temporalmente el archivo
        nombre_archivo = f"{reserva.usuario.nombre}_{reserva.fecha.strftime('%Y-%m-%d')}.pdf"
        filename = secure_filename(nombre_archivo)
        filepath = os.path.join('uploads', filename)
        archivo.save(filepath)

        # Subir a Drive
        archivo_id = subir_archivo_a_drive(
            drive_service,
            filepath,
            reserva.usuario.nombre,
            current_app.config['FOLDER_ID']
        )

        # Guardar el nombre del archivo en la reserva
        reserva.hoja_firmada_nombre = filename
        reserva.hoja_firmada_id = archivo_id
        db.session.commit()

        # Borrar el archivo local si existe
        if os.path.exists(filepath):
            os.remove(filepath)

        flash('Archivo subido correctamente a Google Drive', 'success')
    else:
        flash('No se ha seleccionado ningún archivo', 'danger')

    return redirect(url_for('main.reservar_material'))

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                         RUTAS DE SUSTITUCIONES                         ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/sustituciones")
@login_required
def ver_sustituciones():
    hoy = date.today()
    zona = timezone("Europe/Madrid")
    ahora = datetime.now().time()
    page = request.args.get("page", 1, type=int)

    grupos = Grupo.query.order_by(Grupo.orden).all()
    profesores = Usuario.query.order_by(Usuario.nombre).all()

    if current_user.rol in ["jefatura", "tic"]:
        sustituciones_proximas = Sustitucion.query.filter(
            (Sustitucion.fecha > hoy) |
            ((Sustitucion.fecha == hoy) & (Sustitucion.hora_fin > ahora))
        ).order_by(Sustitucion.fecha.asc(), Sustitucion.hora_inicio.asc()).all()

        sustituciones_pasadas = Sustitucion.query.filter(
            (Sustitucion.fecha < hoy) |
            ((Sustitucion.fecha == hoy) & (Sustitucion.hora_fin <= ahora))
        ).order_by(Sustitucion.fecha.desc(), Sustitucion.hora_inicio.desc()).all()

    else:
        sustituciones_proximas = Sustitucion.query.filter(
            Sustitucion.sustituto_id == current_user.id,
            (Sustitucion.fecha > hoy) |
            ((Sustitucion.fecha == hoy) & (Sustitucion.hora_fin > ahora))
        ).order_by(Sustitucion.fecha.asc(), Sustitucion.hora_inicio.asc()).all()

        ultimas_sustituciones = Sustitucion.query.filter(
            Sustitucion.sustituto_id == current_user.id,
            (Sustitucion.fecha < hoy) |
            ((Sustitucion.fecha == hoy) & (Sustitucion.hora_fin <= ahora))
        ).order_by(Sustitucion.fecha.desc(), Sustitucion.hora_inicio.desc()).limit(10).all()

        sustituciones_pasadas = None  # No se muestran para docentes normales

    return render_template(
        "sustituciones.html",
        grupos=grupos,
        profesores=profesores,
        sustituciones_proximas=sustituciones_proximas,
        ultimas_sustituciones=ultimas_sustituciones if current_user.rol not in ["jefatura", "tic"] else None,
        sustituciones_pasadas=sustituciones_pasadas if current_user.rol in ["jefatura", "tic"] else None
    )

@main_bp.route("/sustituciones/nueva", methods=["POST"])
@login_required
@rol_requerido("jefatura")
def nueva_sustitucion():
    if current_user.rol != "jefatura":
        flash("No tienes permiso para crear sustituciones", "danger")
        return redirect(url_for("main.ver_sustituciones"))

    fecha = request.form.get("fecha")
    hora_inicio = request.form.get("hora_inicio")
    hora_fin = request.form.get("hora_fin")
    grupo_id = request.form.get("grupo_id")
    id_sustituido = request.form.get("sustituido_id")
    id_sustituto = request.form.get("sustituto_id")
    observaciones = request.form.get("observaciones")

    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()

    # Crear carpeta de Drive
    #enlace_material = crear_carpeta_sustitucion(id_sustituido, fecha_dt, hora_inicio, grupo_id)
    sustituido = Usuario.query.get(id_sustituido)
    enlace_material = sustituido.enlace_materiales

    sustitucion = Sustitucion(
        sustituido_id=id_sustituido,
        sustituto_id=id_sustituto,
        fecha=fecha_dt,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
        grupo_id=grupo_id,
        observaciones=observaciones,
        enlace_drive=enlace_material
    )

    db.session.add(sustitucion)
    db.session.commit()

    # Obtener objetos de usuario y grupo para el SMS
    sustituto = Usuario.query.get(id_sustituto)
    sustituido = Usuario.query.get(id_sustituido)
    grupo = Grupo.query.get(grupo_id)

    # Dentro de tu función de envío de SMS
    confirmar_url = url_for("main.confirmar_sustitucion", sustitucion_id=sustitucion.id, _external=True)

    mensaje = (
        f"⚠️ Jefatura de estudios: "
        f"Tienes sustitución en {grupo.nombre}, "
        f"día {fecha_dt.strftime('%d/%m/%Y')} de {hora_inicio} a {hora_fin}. "
        f"✅ Confirma lectura: {confirmar_url}\n"
        f"📂 Material sustitución: {enlace_material}"
    )

    ok, respuesta = enviar_sms_esendex(sustituto.telefono, mensaje)

    if ok:
        flash("Sustitución creada y SMS enviado", "success")
    else:
        flash(f"Sustitución creada, pero error al enviar SMS: {respuesta}", "warning")

    return redirect(url_for("main.ver_sustituciones"))

@main_bp.route("/sustituciones/cancelar/<int:sustitucion_id>", methods=["POST"])
@login_required
@rol_requerido("jefatura")
def cancelar_sustitucion(sustitucion_id):
    sustitucion = Sustitucion.query.get_or_404(sustitucion_id)
    db.session.delete(sustitucion)
    db.session.commit()
    flash("Sustitución cancelada correctamente.", "success")
    return redirect(url_for("main.ver_sustituciones"))

@main_bp.route("/sustituciones/confirmar/<int:sustitucion_id>", methods=["POST", "GET"])
@login_required
def confirmar_sustitucion(sustitucion_id):
    sustitucion = Sustitucion.query.get_or_404(sustitucion_id)
    if current_user.id != sustitucion.sustituto_id:
        flash("No tienes permiso para confirmar esta sustitución.", "danger")
        return redirect(url_for("main.ver_sustituciones"))

    sustitucion.confirmado = True
    db.session.commit()
    flash("Sustitución confirmada correctamente.", "success")
    return redirect(url_for("main.ver_sustituciones"))

@main_bp.route("/sustituciones/historial")
@login_required
def historial_sustituciones():
    sustituciones = Sustitucion.query.order_by(Sustitucion.fecha.desc(), Sustitucion.hora_inicio).all()
    return render_template("historial_sustituciones.html", sustituciones=sustituciones)
# ╔════════════════════════════════════════════════════════════════════════╗
# ║                          RUTAS DE UBICACIONES                          ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route('/ubicaciones')
@login_required
@rol_requerido("tic")
def listado_ubicaciones():
    ubicaciones = Ubicacion.query.order_by(Ubicacion.planta.asc(), Ubicacion.nombre.asc()).all()
    return render_template('ubicaciones.html', ubicaciones=ubicaciones)

@main_bp.route('/actualizar', methods=['POST'])
@login_required
@rol_requerido("tic")
def actualizar_ubicacion():
    data = request.get_json()
    ubicacion_id = data.get('id')
    campo = data.get('campo')
    valor = data.get('valor')

    ubicacion = Ubicacion.query.get(ubicacion_id)
    if ubicacion and hasattr(ubicacion, campo):
        setattr(ubicacion, campo, valor)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 400

@main_bp.route("/ubicaciones/nueva", methods=["POST"])
@login_required
@rol_requerido("tic")
def nueva_ubicacion():

    nueva = Ubicacion(
        tipo=request.form.get("tipo"),
        nombre=request.form.get("nombre"),
        numero_puerta=request.form.get("numero_puerta"),
        planta=request.form.get("planta"),
        observaciones=request.form.get("observaciones")
    )
    db.session.add(nueva)
    db.session.commit()
    return redirect(url_for('main.listado_ubicaciones'))

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                           RUTAS DE EQUIPOS                             ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route('/portatiles')
@login_required
@rol_requerido("tic")
def ver_portatiles():
    filtro = request.args.get('filtro')

    query = Dispositivo.query.filter_by(tipo='portatil')

    if filtro:
        query = query.join(Ubicacion).filter(Ubicacion.nombre == f'CARRO PORTÁTILES {filtro}')

    portatiles = query.order_by(Dispositivo.etiqueta).all()
    ubicaciones = Ubicacion.query.order_by(Ubicacion.planta, Ubicacion.numero_puerta).all()

    return render_template('material/portatiles.html', vista_actual='portatiles', portatiles=portatiles, filtro=filtro, ubicaciones=ubicaciones)

@main_bp.route('/dispositivos/actualizar-multiple', methods=['POST'])
@login_required
@rol_requerido("tic")
def actualizar_multiples_dispositivos():
    data = request.json  # Lista de objetos con datos de dispositivos
    try:
        for item in data:
            dispositivo = Dispositivo.query.get(item.get('id'))
            if not dispositivo:
                continue
            for campo in ['etiqueta', 'numero_serie', 'estado', 'ubicacion_id', 'marca', 'modelo', 'observaciones']:
                if campo in item:
                    setattr(dispositivo, campo, item[campo] or None)
        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e))

@main_bp.route('/sobremesa')
@login_required
@rol_requerido("tic")
def ver_sobremesa():
    filtro = request.args.get('filtro')

    query = Dispositivo.query.options(joinedload(Dispositivo.ubicacion)).filter_by(tipo='sobremesa')

    if filtro:
        query = query.join(Ubicacion).filter(Ubicacion.nombre == f'SOBREMESA {filtro}')

    sobremesas = query.order_by(Dispositivo.etiqueta).all()
    ubicaciones = Ubicacion.query.order_by(Ubicacion.nombre).all()

    return render_template('material/sobremesa.html', vista_actual='sobremesas', sobremesas=sobremesas, filtro=filtro, ubicaciones=ubicaciones)

@main_bp.route('/allinone')
@login_required
@rol_requerido("tic")
def ver_allinone():
    allinones = Dispositivo.query.filter_by(tipo='allinone').order_by(Dispositivo.nombre).all()
    ubicaciones = Ubicacion.query.order_by(Ubicacion.nombre).all()
    return render_template('material/allinone.html', vista_actual='allinones', allinones=allinones, ubicaciones=ubicaciones)

@main_bp.route('/tablets')
@login_required
@rol_requerido("tic")
def ver_tablets():
    filtro = request.args.get('filtro')
    query = Dispositivo.query.options(joinedload(Dispositivo.ubicacion)).filter_by(tipo='tablet')

    if filtro == '1':
        query = query.join(Ubicacion).filter(Ubicacion.nombre == 'CARRO TABLETS 1')
    elif filtro == '2':
        query = query.join(Ubicacion).filter(Ubicacion.nombre == 'CARRO TABLETS 2')

    tablets = query.order_by(Dispositivo.nombre).all()
    ubicaciones = Ubicacion.query.order_by(Ubicacion.nombre).all()
    return render_template('material/tablets.html', vista_actual='tablets', tablets=tablets, ubicaciones=ubicaciones, filtro=filtro)

@main_bp.route('/pizarras')
@login_required
@rol_requerido("tic")
def ver_pizarras():
    query = Dispositivo.query.filter_by(tipo='pizarra')

    pizarras = query.order_by(Dispositivo.id).all()
    ubicaciones = Ubicacion.query.order_by(Ubicacion.nombre).all()
    return render_template("material/pizarras.html", vista_actual='pizarras', pizarras=pizarras, ubicaciones=ubicaciones)

@main_bp.route('/inventario')
@login_required
@rol_requerido("tic")
def ver_inventario():
    dispositivos = Dispositivo.query.options(joinedload(Dispositivo.ubicacion)).all()

    tipos_raw = db.session.query(Dispositivo.tipo).distinct().all()
    estados_raw = db.session.query(Dispositivo.estado).distinct().all()

    tipos = sorted([t[0] for t in tipos_raw if t[0]])
    estados = sorted([e[0] for e in estados_raw if e[0]])

    return render_template(
        'material/inventario.html',
        dispositivos=dispositivos,
        tipos=tipos,
        estados=estados,
        vista_actual='inventario'
    )

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                            RUTAS DE RESERVAS                           ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route("/asignar-dispositivos/<int:reserva_id>", methods=["GET", "POST"])
@login_required
def asignar_dispositivos(reserva_id):
    # Diccionario para filtros visuales
    CARROS = {
        "p1": 39,  # Carro portátiles 1
        "p2": 40,
        "p3": 43,
        "t1": 41,  # Carro tablets 1
        "t2": 42
    }

    # Listas de IDs de ubicación por tipo de equipo
    CARROS_PORTATILES = [39, 40, 43]
    CARROS_TABLETS = [41, 42]

    reserva = ReservaInformatica.query.get_or_404(reserva_id)
    filtro = request.args.get("filtro")

    # Determinar los carros válidos según tipo de equipo
    if reserva.tipo_equipo.lower() == "portátil":
        ubicaciones_validas = CARROS_PORTATILES
    elif reserva.tipo_equipo.lower() == "tablet":
        ubicaciones_validas = CARROS_TABLETS
    else:
        ubicaciones_validas = []

    # Construir consulta filtrando por tipo y ubicaciones válidas
    query = Dispositivo.query.filter(
        Dispositivo.tipo == reserva.tipo_equipo,
        Dispositivo.estado == 'activo',
        Dispositivo.ubicacion_id.in_(ubicaciones_validas)
    )

    # Aplicar filtro por carro específico si se indica (como 'p1', 't2', etc.)
    if filtro in CARROS:
        query = query.filter(Dispositivo.ubicacion_id == CARROS[filtro])

    dispositivos = query.all()

    # Buscar dispositivos ocupados en otras reservas en la misma fecha y franja
    reservas_misma_hora = ReservaInformatica.query.filter(
        ReservaInformatica.fecha == reserva.fecha,
        ReservaInformatica.franja_horaria == reserva.franja_horaria,
        ReservaInformatica.estado == "ACEPTADA",
        ReservaInformatica.id != reserva.id
    ).all()

    dispositivos_ocupados = set()
    for r in reservas_misma_hora:
        dispositivos_ocupados.update([d.id for d in r.dispositivos])

    # Procesar asignación si es POST
    if request.method == "POST":
        ids = request.form.getlist("dispositivos")
        seleccionados = Dispositivo.query.filter(Dispositivo.id.in_(ids)).all()
        reserva.dispositivos = seleccionados
        reserva.estado = "ACEPTADA"
        db.session.commit()
        flash("Reserva aceptada y dispositivos asignados.", "success")
        return redirect(url_for("main.reservas_material"))

    return render_template("reserva_material/asignar_dispositivos.html",
                           reserva=reserva,
                           dispositivos=dispositivos,
                           dispositivos_ocupados=dispositivos_ocupados,
                           filtro=filtro,
                           carros=CARROS)

@main_bp.route("/denegar-reserva", methods=["POST"])
@login_required
def denegar_reserva():
    if current_user.rol not in ["tic", "jefatura"]:
        abort(403)
    reserva_id = request.form.get("reserva_id")
    reserva = ReservaInformatica.query.get_or_404(reserva_id)
    reserva.estado = "DENEGADA"
    db.session.commit()
    flash("Reserva denegada correctamente.", "info")
    return redirect(url_for("main.reservas_material"))

@main_bp.route('/descargar-hoja/<int:reserva_id>')
@login_required
def descargar_hoja(reserva_id):
    reserva = ReservaInformatica.query.get_or_404(reserva_id)

    if not reserva.hoja_firmada_id:
        flash("No hay hoja firmada disponible para esta reserva", "danger")
        return redirect(url_for('main.reservar_material'))

    drive_service = build_drive_service()

    try:
        # Obtener enlace de descarga (webContentLink solo si archivo es público)
        file = drive_service.files().get(
            fileId=reserva.hoja_firmada_id,
            fields='webContentLink',
            supportsAllDrives=True
        ).execute()

        return redirect(file['webContentLink'])

    except Exception as e:
        flash("No se pudo descargar el archivo desde Google Drive", "danger")
        return redirect(url_for('main.reservar_material'))

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                          RUTAS DE INCIDENCIAS                          ║
# ╚════════════════════════════════════════════════════════════════════════╝

@main_bp.route('/incidencias/nueva', methods=['GET', 'POST'])
@login_required
def nueva_incidencia():
    es_tic = current_user.rol == 'tic'
    if request.method == 'POST':
        ubicacion = request.form['ubicacion']
        equipo = request.form.get('equipo_id')  if es_tic else None
        descripcion = request.form['descripcion']

        nueva = Incidencia(
            ubicacion=ubicacion,
            equipo_id=equipo,
            descripcion=descripcion,
            docente_id=current_user.id
        )
        db.session.add(nueva)
        db.session.commit()
        flash('Incidencia registrada correctamente.', 'success')
        return redirect(url_for('main.mostrar_incidencias'))

    equipos = Dispositivo.query.all()
    equipos_json = [
        {
            'id': e.id,
            'nombre': e.nombre,
            'tipo': e.tipo,
            'ubicacion': e.ubicacion.nombre if e.ubicacion else 'Sin ubicación'
        }
        for e in equipos
    ]

    return render_template('incidencias/nueva.html',
                           es_tic=es_tic,
                           equipos_json=equipos_json)

@main_bp.route('/incidencias/<int:id>/cancelar', methods=['POST'])
@login_required
def cancelar_incidencia(id):
    incidencia = Incidencia.query.get_or_404(id)
    if incidencia.docente_id != current_user.id:
        abort(403)
    incidencia.estado = 'Cancelada'
    db.session.commit()
    flash('Incidencia cancelada.', 'warning')
    return redirect(url_for('main.mostrar_incidencias'))

@main_bp.route('/incidencias')
@login_required
def mostrar_incidencias():
    if current_user.rol == 'tic':
        incidencias = Incidencia.query.order_by(Incidencia.fecha_hora.desc()).all()
    else:
        incidencias = Incidencia.query.filter_by(docente_id=current_user.id).order_by(Incidencia.fecha_hora.desc()).all()

    # Crear diccionario con id y nombre de todos los usuarios
    usuarios = Usuario.query.all()
    usuarios_dict = {u.id: u.nombre for u in usuarios}

    return render_template('incidencias/listado.html',
                           incidencias=incidencias,
                           es_tic=(current_user.rol == 'tic'),
                           usuarios_dict=usuarios_dict)

@main_bp.route('/incidencias/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_incidencia(id):
    incidencia = Incidencia.query.get_or_404(id)
    es_tic = current_user.rol == 'tic'

    if current_user.id != incidencia.docente_id and not es_tic:
        abort(403)

    # Obtener el equipo actual (por id)
    equipo_actual = incidencia.equipo
    tipo_equipo_actual = equipo_actual.tipo if equipo_actual else ""

    estado_anterior = incidencia.estado

    if request.method == 'POST':
        incidencia.ubicacion = request.form['ubicacion']
        equipo_id_str = request.form['equipo']
        incidencia.equipo_id = int(equipo_id_str) if equipo_id_str else None
        incidencia.descripcion = request.form['descripcion']

        if es_tic:
            incidencia.estado = request.form['estado']
            incidencia.prioridad = request.form['prioridad']

        db.session.commit()

        return redirect(url_for('main.mostrar_incidencias'))


    # Obtener todos los equipos y pasarlos en formato JSON
    equipos = Dispositivo.query.all()
    equipos_json = [
        {'id': e.id, 'nombre': e.nombre, 'ubicacion': e.ubicacion, 'tipo': e.tipo}
        for e in equipos
    ]
    print(equipos_json)

    return render_template('incidencias/editar.html',
                           incidencia=incidencia,
                           es_tic=es_tic,
                           tipo_equipo_actual=tipo_equipo_actual,
                           equipos_json=equipos_json)

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                         RUTAS DE MANTENIMIENTO                         ║
# ╚════════════════════════════════════════════════════════════════════════╝
@main_bp.route('/mantenimiento')
@login_required
def listar_incidencias_mantenimiento():
    if current_user.rol in ['tic', 'jefatura']:
        pendientes = IncidenciaMantenimiento.query.filter(IncidenciaMantenimiento.estado != 'Resuelta').order_by(IncidenciaMantenimiento.fecha_hora.desc()).all()
        resueltas = IncidenciaMantenimiento.query.filter_by(estado='Resuelta').order_by(IncidenciaMantenimiento.fecha_hora.desc()).paginate(per_page=10)
    else:
        pendientes = IncidenciaMantenimiento.query.filter_by(docente_id=current_user.id).filter(IncidenciaMantenimiento.estado != 'Resuelta').order_by(IncidenciaMantenimiento.fecha_hora.desc()).all()
        resueltas = IncidenciaMantenimiento.query.filter_by(docente_id=current_user.id, estado='Resuelta').order_by(IncidenciaMantenimiento.fecha_hora.desc()).paginate(per_page=10)

    usuarios = Usuario.query.all()
    usuarios_dict = {u.id: u.nombre for u in usuarios}

    return render_template('mantenimiento/listado.html',
                           pendientes=pendientes,
                           resueltas=resueltas,
                           usuarios_dict=usuarios_dict,
                           es_admin=(current_user.rol in ['tic', 'jefatura']))

@main_bp.route('/mantenimiento/nueva', methods=['GET', 'POST'])
@login_required
def nueva_incidencia_mantenimiento():
    es_admin = current_user.rol in ['tic', 'jefatura']

    if request.method == 'POST':
        incidencia = IncidenciaMantenimiento(
            ubicacion=request.form['ubicacion'],
            descripcion=request.form['descripcion'],
            docente_id=current_user.id
        )

        if es_admin:
            incidencia.estado = request.form['estado']
            incidencia.prioridad = request.form['prioridad']

        db.session.add(incidencia)
        db.session.commit()
        flash('Incidencia registrada correctamente.', 'success')
        return redirect(url_for('main.listar_incidencias_mantenimiento'))

    return render_template('mantenimiento/nueva.html', es_admin=es_admin)

@main_bp.route('/mantenimiento/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_incidencia_mantenimiento(id):
    incidencia = IncidenciaMantenimiento.query.get_or_404(id)
    es_admin = current_user.rol in ['tic', 'jefatura']

    if current_user.id != incidencia.docente_id and not es_admin:
        abort(403)

    if request.method == 'POST':
        incidencia.ubicacion = request.form['ubicacion']
        incidencia.descripcion = request.form['descripcion']

        if es_admin:
            nuevo_estado = request.form['estado']
            incidencia.estado = nuevo_estado
            incidencia.prioridad = request.form['prioridad']

            # Guardar la fecha de resolución si se marca como resuelta
            if nuevo_estado == 'Resuelta' and not incidencia.fecha_resolucion:
                incidencia.fecha_resolucion = datetime.now()

        db.session.commit()
        flash('Incidencia actualizada correctamente.', 'success')
        return redirect(url_for('main.listar_incidencias_mantenimiento'))

    return render_template('mantenimiento/editar.html',
                           incidencia=incidencia,
                           es_admin=es_admin)

@main_bp.route('/mantenimiento/pdf')
@login_required
def descargar_pdf_mantenimiento():
    if current_user.rol not in ['tic', 'jefatura']:
        abort(403)

    incidencias = IncidenciaMantenimiento.query.filter(IncidenciaMantenimiento.estado != 'Resuelta').order_by(IncidenciaMantenimiento.fecha_hora.desc()).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Listado de incidencias de mantenimiento pendientes", styles['Title']))
    elements.append(Spacer(1, 12))

    for inc in incidencias:
        texto = f"<b>Ubicación:</b> {inc.ubicacion}<br/><b>Descripción:</b> {inc.descripcion}<br/><b>Prioridad:</b> {inc.prioridad}<br/><b>Estado:</b> {inc.estado}<br/><br/>"
        elements.append(Paragraph(texto, styles['Normal']))
        elements.append(Spacer(1, 8))

    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="incidencias_mantenimiento.pdf", mimetype='application/pdf')

@main_bp.route('/mantenimiento/<int:id>/cancelar', methods=['POST'])
@login_required
def cancelar_incidencia_mantenimiento(id):
    incidencia = IncidenciaMantenimiento.query.get_or_404(id)
    es_admin = current_user.rol in ['tic', 'jefatura']

    # Solo puede cancelar el creador o un admin (TIC o jefatura)
    if current_user.id != incidencia.docente_id and not es_admin:
        abort(403)

    incidencia.estado = 'Cancelada'
    db.session.commit()
    flash('Incidencia cancelada correctamente.', 'success')
    return redirect(url_for('main.listar_incidencias_mantenimiento'))

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                          RUTAS DE ABSENTISMO                           ║
# ╚════════════════════════════════════════════════════════════════════════╝

def calcular_anio_academico(mes):
    """Devuelve el año académico al que pertenece un mes del curso."""
    mes = mes.lower()
    ahora = datetime.now()
    if mes in ["septiembre", "octubre", "noviembre", "diciembre"]:
        return 2025
    else:
        return 2026

@main_bp.route("/absentismo")
@login_required
def ver_absentismo():

    meses_validos = ["septiembre", "octubre", "noviembre", "diciembre",
                     "enero", "febrero", "marzo", "abril", "mayo", "junio"]

    mes_actual = datetime.now().strftime("%B").lower()
    if mes_actual not in meses_validos:
        mes_actual = "septiembre"

    return redirect(url_for("main.ver_absentismo_mes", mes=mes_actual))

@main_bp.route('/absentismo/<mes>')
@login_required
def ver_absentismo_mes(mes):
    from datetime import datetime

    # Año académico real, dependiendo del mes
    anio_actual = calcular_anio_academico(mes)

    todos_los_grupos = Grupo.query.order_by(Grupo.orden).all()

    meses = ["septiembre", "octubre", "noviembre", "diciembre",
             "enero", "febrero", "marzo", "abril", "mayo", "junio"]

    informes = (
        InformeFaltas.query
        .options(selectinload(InformeFaltas.alumnos))
        .filter_by(mes=mes, anio=anio_actual)
        .all()
    )

    informes_por_grupo = {i.grupo_id: i for i in informes}
    print("informes_por_grupo:", informes_por_grupo)
    for grupo_id, informe in informes_por_grupo.items():
        print(f"Grupo ID: {grupo_id}")
        print(f"Informe ID: {informe.id}")
        print(f"Número de alumnos en informe: {len(informe.alumnos)}")
        for ia in informe.alumnos:
            print(
                f"  Alumno ID: {ia.alumno_id}, CJ: {ia.faltas_justificadas}, CI: {ia.faltas_injustificadas}, %CI: {ia.porcentaje_injustificadas}")


    return render_template(
            'absentismo/listado_absentismo.html',
            mes_actual=mes,
            anio_actual=anio_actual,
            grupos=todos_los_grupos,
            informes_por_grupo=informes_por_grupo,
            meses=meses
        )

def normalizar(texto):
    if not texto:
        return ''
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto.lower().strip()

@main_bp.route('/subir-informe/<int:grupo_id>/<mes>', methods=['POST'])
@login_required
def subir_informe(grupo_id, mes):
    grupo = Grupo.query.get_or_404(grupo_id)
    anio = 2025 if mes in ['septiembre', 'octubre', 'noviembre', 'diciembre'] else 2026

    archivo = request.files.get('archivo')
    if not archivo or archivo.filename == '':
        flash("Debes seleccionar un archivo CSV", "danger")
        return redirect(url_for('main.ver_absentismo_mes', mes=mes))

    filename = secure_filename(archivo.filename)
    carpeta_destino = Path(current_app.config['UPLOAD_FOLDER']) / str(anio) / mes
    # Genera ruta completa
    ruta_completa = carpeta_destino / filename

    archivo.save(str(ruta_completa))

    # Intentar leer CSV
    try:
        df = pd.read_csv(ruta_completa, encoding='utf-8', sep=',', dtype=str)
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(ruta_completa, encoding='latin1', sep=',', dtype=str)
        except Exception:
            flash("Error al leer el archivo. Asegúrate de que es un CSV separado por ','", "danger")
            return redirect(url_for('main.ver_absentismo_mes', mes=mes))
    except Exception:
        flash("Error al leer el archivo. Asegúrate de que es un CSV separado por ','", "danger")
        return redirect(url_for('main.ver_absentismo_mes', mes=mes))

    # Normalizar cabeceras si vienen mal
    df.columns = [col.strip().strip('"') for col in df.columns]

    # Comprobar que existe la columna "Alumno/a"
    if "Alumno/a" not in df.columns:
        print("El archivo no contiene una columna 'Alumno/a'")
        flash("El archivo no contiene una columna 'Alumno/a'", "danger")
        return redirect(url_for('main.ver_absentismo_mes', mes=mes))

    # Crear o recuperar informe
    informe = InformeFaltas.query.filter_by(grupo_id=grupo.id, mes=mes, anio=anio).first()
    if not informe:
        informe = InformeFaltas(grupo_id=grupo.id, mes=mes, anio=anio, archivo_csv=filename)
        db.session.add(informe)
        db.session.commit()

    # Borrar registros anteriores del informe
    InformeAlumno.query.filter_by(informe_id=informe.id).delete()

    # Crear diccionario de alumnos del grupo
    alumnos_dict = {
        (normalizar(a.nombre), normalizar(a.apellidos)): a
        for a in Alumno.query.filter_by(grupo_id=grupo.id).all()
    }

    registros_insertados = 0

    columnas_fecha = df.columns[1:]
    dias_lectivos_totales = len(columnas_fecha)

    for _, fila in df.iterrows():
        nombre_completo = fila.get("Alumno/a", "").strip()
        if ',' not in nombre_completo:
            continue

        # Separar "Apellidos, Nombre"
        apellidos_csv, nombre_csv = [parte.strip() for parte in nombre_completo.split(',', 1)]
        clave = (normalizar(nombre_csv), normalizar(apellidos_csv))

        alumno = alumnos_dict.get(clave)
        if not alumno:
            continue  # Alumno no encontrado en este grupo

        # Tomar solo las columnas de fechas
        valores = fila[columnas_fecha]

        # Normalizar: mayúsculas + quitar espacios
        valores = [str(v).strip().upper() for v in valores]

        # Contar tipos de faltas
        cj = valores.count('CJ')
        ci = valores.count('CI')

        # Calcular % CI sobre días lectivos
        porcentaje = (ci / dias_lectivos_totales * 100) if dias_lectivos_totales > 0 else 0
        absentista = porcentaje >= 20

        db.session.add(InformeAlumno(
            informe_id=informe.id,
            alumno_id=alumno.id,
            faltas_justificadas=cj,
            faltas_injustificadas=ci,
            porcentaje_injustificadas=porcentaje,
            absentista=absentista
        ))
        registros_insertados += 1

    db.session.commit()
    db.session.refresh(informe)

    print("registros_insertados:", registros_insertados)

    if registros_insertados == 0:
        flash("Archivo subido pero no se han encontrado coincidencias de alumnos", "warning")
    else:
        flash(f"Informe subido correctamente. {registros_insertados} alumnos procesados.", "success")

    return redirect(url_for("main.ver_absentismo_mes", mes=mes))

@main_bp.route('/absentismo/<mes>/grupo/<int:grupo_id>')
@login_required
def ver_informe_grupo(mes, grupo_id):
    anio = calcular_anio_academico(mes)
    grupo = Grupo.query.get_or_404(grupo_id)
    informe = (
        InformeFaltas.query
        .options(selectinload(InformeFaltas.alumnos).selectinload(InformeAlumno.alumno))
        .filter_by(grupo_id=grupo_id, mes=mes, anio=anio)
        .first()
    )
    if not informe:
        flash("No se encontró el informe para este grupo y mes.", "warning")
        return redirect(url_for('main.ver_absentismo_mes', mes=mes))

    return render_template(
        'absentismo/informe_grupo.html',
        grupo=grupo,
        mes=mes,
        anio=anio,
        informe=informe
    )

@main_bp.route('/descargar-informe/<int:grupo_id>/<mes>')
@login_required
def descargar_informe(grupo_id, mes):
    anio_actual = calcular_anio_academico(mes)

    informe = InformeFaltas.query.filter_by(grupo_id=grupo_id, mes=mes, anio=anio_actual).first()

    if not informe:
        flash("No se encontró el informe solicitado.", "warning")
        return redirect(url_for("main.ver_absentismo_mes", mes=mes))

    # Crear un Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de absentismo"

    # Establecer tamaño de fuente para toda la hoja
    fuente_general = Font(size=12)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # desde la segunda fila, saltamos encabezado
        for cell in row:
            cell.font = fuente_general

    # Encabezados
    encabezados = ["Alumno", "Faltas justificadas", "Faltas injustificadas", "% Injustificadas"]
    ws.append(encabezados)

    # Estilo para absentistas
    fill_amarillo = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    for entrada in informe.alumnos:
        # Añadir fila con nombre completo y valores
        ws.append([
            f"{entrada.alumno.nombre} {entrada.alumno.apellidos}",
            entrada.faltas_justificadas,
            entrada.faltas_injustificadas,
            None  # dejamos temporalmente la celda del porcentaje en blanco
        ])

        # Obtenemos la fila actual
        fila_actual = ws.max_row

        # Insertamos el porcentaje en formato correcto
        celda_porcentaje = ws.cell(row=fila_actual, column=4)
        celda_porcentaje.value = round(entrada.porcentaje_injustificadas / 100, 4)  # 10 -> 0.10
        celda_porcentaje.number_format = '0.0%'  # Formato de porcentaje en Excel

        # Si es absentista, rellenamos toda la fila con color
        if entrada.absentista:
            for cell in ws[fila_actual]:
                cell.fill = fill_amarillo

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_grupo = informe.grupo.nombre.replace(" ", "_")
    filename = f"informe_{nombre_grupo}_{mes}_{anio_actual}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@main_bp.route('/borrar-informe/<int:grupo_id>/<mes>', methods=['POST'])
@login_required
def borrar_informe(grupo_id, mes):
    grupo = Grupo.query.get_or_404(grupo_id)
    anio = 2025 if mes in ['septiembre', 'octubre', 'noviembre', 'diciembre'] else 2026

    informe = InformeFaltas.query.filter_by(grupo_id=grupo.id, mes=mes, anio=anio).first()

    if not informe:
        flash("No hay informe para este grupo y mes", "warning")
        return redirect(url_for("main.ver_absentismo_mes", mes=mes))

    InformeAlumno.query.filter_by(informe_id=informe.id).delete()
    db.session.delete(informe)
    db.session.commit()

    flash("Informe eliminado correctamente", "success")
    return redirect(url_for("main.ver_absentismo_mes", mes=mes))

@main_bp.route('/descargar-informe-horizontal/<int:grupo_id>')
@login_required
def descargar_informe_horizontal(grupo_id):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from io import BytesIO

    grupo = Grupo.query.get_or_404(grupo_id)

    # Determinar el curso académico actual (puedes ajustar esto)
    mes_actual = datetime.now().strftime('%B').lower()
    anio_actual = datetime.now().year
    anio_inicio = calcular_anio_academico(mes_actual)
    anio_academico = f"{anio_inicio}-{anio_inicio + 1}"

    # Buscar todos los informes de ese grupo en ese curso académico
    informes = InformeFaltas.query.filter_by(grupo_id=grupo_id, anio=anio_inicio).all()

    if not informes:
        flash("No hay informes para este grupo en este curso.", "warning")
        return redirect(url_for("main.ver_absentismo_mes", mes=mes_actual))

    # Crear hoja Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{grupo.nombre} {anio_academico}"

    # Cabecera dinámica
    cabecera = ["Alumno"]
    meses_ordenados = ["septiembre", "octubre", "noviembre", "diciembre", "enero", "febrero", "marzo", "abril", "mayo", "junio"]
    informes_ordenados = sorted(informes, key=lambda i: meses_ordenados.index(i.mes))

    for inf in informes_ordenados:
        cabecera += [f"CJ {inf.mes[:3]}", f"CI {inf.mes[:3]}", f"% {inf.mes[:3]}"]

    ws.append(cabecera)

    # Estilo para absentistas
    fill_amarillo = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    font_bold = Font(bold=True)

    # Alumnos
    todos_alumnos = Alumno.query.filter_by(grupo_id=grupo_id).order_by(Alumno.apellidos).all()

    for alumno in todos_alumnos:
        fila = [f"{alumno.nombre} {alumno.apellidos}"]

        for inf in informes_ordenados:
            entrada = next((e for e in inf.alumnos if e.alumno_id == alumno.id), None)
            if entrada:
                fila += [
                    entrada.faltas_justificadas,
                    entrada.faltas_injustificadas,
                    round(entrada.porcentaje_injustificadas, 1)
                ]
            else:
                fila += ["", "", ""]

        ws.append(fila)

        # Pintar de amarillo si algún mes fue absentista
        row = ws.max_row
        for i, inf in enumerate(informes_ordenados):
            entrada = next((e for e in inf.alumnos if e.alumno_id == alumno.id), None)
            if entrada and entrada.absentista:
                for col in range(2 + i * 3, 5 + i * 3):  # columnas CJ, CI, %
                    ws.cell(row=row, column=col).fill = fill_amarillo

    # Ajustar ancho
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"informe_acumulado_{grupo.nombre.replace(' ', '_')}_{anio_academico}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@main_bp.route('/absentismo/resumen')
@login_required
def resumen_absentismo():
    # Determinar año académico actual
    mes_actual = datetime.now().strftime("%B").lower()
    anio_inicio = calcular_anio_academico(mes_actual)
    anio_academico = f"{anio_inicio}-{anio_inicio + 1}"

    # Obtener todos los grupos
    grupos = Grupo.query.order_by(Grupo.nombre).all()

    resumen_grupos = []

    for grupo in grupos:
        informes = InformeFaltas.query.filter_by(grupo_id=grupo.id, anio=anio_inicio).all()
        total_alumnos = Alumno.query.filter_by(grupo_id=grupo.id).count()
        total_informes = len(informes)

        total_absentistas = 0
        total_porcentajes = 0
        total_entradas = 0

        for inf in informes:
            for entrada in inf.alumnos:
                total_porcentajes += entrada.porcentaje_injustificadas
                total_entradas += 1
                if entrada.absentista:
                    total_absentistas += 1

        resumen_grupos.append({
            "grupo": grupo,
            "num_meses": total_informes,
            "num_alumnos": total_alumnos,
            "num_absentistas": total_absentistas,
            "media_porcentaje": round(total_porcentajes / total_entradas, 1) if total_entradas else 0
        })

    return render_template("absentismo/resumen.html",
                           anio_academico=anio_academico,
                           resumen_grupos=resumen_grupos)